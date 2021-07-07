from openpyxl import load_workbook
from array import array
import numpy as np


def coordinates_to_float(val):
    global low_lat, high_lat, low_lon, high_lon
    low_lat = float(val['A' + i].value)
    high_lat = float(val['B' + i].value)
    low_lon = float(val['C' + i].value)
    high_lon = float(val['D' + i].value)


def soil_range():
    global low_soil, high_soil
    low_soil = float(soil_sheet_val['E' + i].value)
    high_soil = float(soil_sheet_val['F' + i].value)


def temperature_range():
    global low_temp, high_temp
    low_temp = int(temp_sheet_val['E' + i].value)
    high_temp = int(temp_sheet_val['F' + i].value)


wb_val = load_workbook(filename='Plants1.1.xlsx')
sheet_val = wb_val['Information about plants']
a_form = sheet_val['A1'].value
info = []
i = '1'
while a_form is not None:
    info.append(a_form)
    i = int(i)
    i += 1
    i = str(i)
    a_form = sheet_val['A' + i].value
flag = False
count = 0
print(info)
temp_doc = load_workbook(filename='Climate.xlsx')
temp_sheet_val = temp_doc['CritTemp']
soil_doc = load_workbook(filename='SoilAcidity.xlsx')
soil_sheet_val = soil_doc['Soil']
lat = input("Enter the latitude of your location: ")
lon = input("Enter the longitude of your location: ")
lat = float(lat)
lon = float(lon)
a_form = temp_sheet_val['A1'].value
i = '1'
low_temp = high_temp = 0
low_lat = high_lat = low_lon = high_lon = 0
low_soil = high_soil = 0
while a_form is not None:
    coordinates_to_float(temp_sheet_val)
    if low_lat <= lat < high_lat:
        if low_lon <= lon < high_lon:
            temperature_range()
    i = int(i)
    i += 1
    i = str(i)
    a_form = temp_sheet_val['A' + i].value
i = '1'
b_form = soil_sheet_val['A' + i].value
while b_form is not None:
    coordinates_to_float(soil_sheet_val)
    if low_lat <= lat < high_lat:
        if low_lon <= lon < high_lon:
            soil_range()
    i = int(i)
    i += 1
    i = str(i)
    b_form = soil_sheet_val['A' + i].value
print("This is the temperatures: from ", low_temp, " to", high_temp)
print("This is the soil acidity: from ", low_soil, " to", high_soil)
print("Enter the name of the desired plant:")
while not flag:
    plant = input()
    for elem in info:
        a_form = elem
        count += 1
        if a_form == plant:
            flag = True
            break
    if not flag:
        print("No plants with that name found. Try again:")
i = str(count)
letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
print("Information about this plant:")
print("Sort: ", sheet_val['B' + i].value)
print("Category: ", sheet_val['C' + i].value)
print("Sun Requirements: ", sheet_val['D' + i].value)
print("Water Preferences: ", sheet_val['E' + i].value)
print("Soil pH requirements: from ", sheet_val['F' + i].value, " to ", sheet_val['G' + i].value)
print("Optimal temperature for")
print("seed swelling: from ", sheet_val['H' + i].value, " to ", sheet_val['I' + i].value)
print("seed germination: from ", sheet_val['J' + i].value, " to ", sheet_val['K' + i].value)
print("laying fruit: from ", sheet_val['L' + i].value, " to ", sheet_val['M' + i].value)
print("Critical temperature for")
print("seedlings: from ", sheet_val['N' + i].value, " to ", sheet_val['O' + i].value)
print("plants: from ", sheet_val['P' + i].value, " to ", sheet_val['Q' + i].value)


def sigmoid(x):
    return 1 / (1 + np.exp(-x))


def derivative(x):
    return x*(1 - x)


def average_value(x, y):
    return (x+y)/2


my_low_soil = float(sheet_val['F' + i].value)
my_high_soil = float(sheet_val['G' + i].value)
low_temp_for_fruits = float(sheet_val['L' + i].value)
high_temp_for_fruits = float(sheet_val['M' + i].value)
graph_doc = load_workbook(filename='Graf.xlsx')
graph_list = graph_doc['List1']
print("This is the average_my_soil: ", average_value(my_low_soil, my_high_soil))
print("This is the average_soil: ", average_value(low_soil, high_soil))
print("This is my high temperature: ", high_temp)
print("This is the average temperature for fruits: ", average_value(low_temp_for_fruits, high_temp_for_fruits)) #(average_ph - 5.5, temp - 23)
x = np.array([[0.7, 0.9, 1.0, 2.0],
              [0.3, 1.1, 4.0, 4.5],
              [1.2, 1.1, -3.0, 2.0],
              [0.6, 0.5, 2.5, 2.0],
              [0.1, 0.1, -1.0, 0.0],
              [0.8, 0.2, 0.5, 3.5],
              [0.3, 0.5, 2.0, 0.5]])
y = np.array([[1, 0, 0, 1, 1, 0, 1]]).T
syn0 = 2*np.random.random((4, 7)) - 1
syn1 = 2*np.random.random((7, 3)) - 1
syn2 = 2*np.random.random((3, 1)) - 1
j = 1
graph_list['A' + str(j)] = "Ошибка прогнозирования"
graph_list['B' + str(j)] = "Кол-во пройденных итераций в тысячах"
for i in range(50000):
    l1 = sigmoid(np.dot(x, syn0))
    l2 = sigmoid(np.dot(l1, syn1))
    l3 = sigmoid(np.dot(l2, syn2))
    l3_delta = (y - l3) * derivative(l3)
    l2_delta = l3_delta.dot(syn2.T) * derivative(l2)
    l1_delta = l2_delta.dot(syn1.T) * derivative(l1)
    syn2 += l2.T.dot(l3_delta)
    syn1 += l1.T.dot(l2_delta)
    syn0 += x.T.dot(l1_delta)
    if (i % 1000) == 0:
        j += 1
        print("Error:" + str(np.mean(np.abs(y - l3))))
        graph_list['A' + str(j)] = ((1000000 * np.mean(np.abs(y - l3))) // 100) / 100
        graph_list['B' + str(j)] = i // 1000
my_data = np.array([[average_value(low_soil, high_soil) - 5.5,
                     average_value(my_low_soil, my_high_soil) - 5.5,
                     high_temp - 23,
                     average_value(low_temp_for_fruits, high_temp_for_fruits) - 23]])
l1 = sigmoid(np.dot(my_data, syn0))
l2 = sigmoid(np.dot(l1, syn1))
l3 = sigmoid(np.dot(l2, syn2))
print("l3: ", l3)
graph_doc.save('Graphs.xlsx')
# 55.6417 37.6728
